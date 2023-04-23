#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jun 23 15:29:18 2020

@author: ex
"""
# Extract the Hirshfeld charges and calculate the condensed dual descriptor, Fukui and Anderson index using the Gaussian file of a molecule with a specific structure but with charges of 0, -1, +1. For gaussian 09 files

import re,os
from openpyxl import Workbook,load_workbook
import numpy
    
entrada_teclado_2=False
archivos_malos=[]

#----------------------------------------------------------------------------#
# Entrada de datos

while entrada_teclado_2 != True:
    opcion=input('1) Files name\n2) All the files in the current file\n3) Full path\nR: ')
    if opcion == '1' or opcion == '2' or opcion=='3':
        entrada_teclado_2=True
    else:
        print('\nInput a correct value')
        
if opcion=='1' :
    rutas=os.getcwd()
    rutas=rutas.split(',')
    archivos=input('Enter the name of the files separated by commas:\n')
    orden_reactividad=int(input('Enter until which order of reactivity you want to calculate: '))
    archivos+=','
    lista_archivos=[]
    nombre=''
    for cadena in range(len(archivos)):
        if archivos[cadena] != ',':
            nombre+=archivos[cadena]
        else:
            lista_archivos.append(nombre)
            nombre=''

if opcion=='2':
    orden_reactividad=int(input('Enter until which order of reactivity you want to calculate: '))
    rutas=os.getcwd()
    rutas=[rutas]

if opcion=='3':
    rutas=input('Rutas: ')
    orden_reactividad=int(input('Enter until which order of reactivity you want to calculate: '))
    rutas=rutas.split(',')

#----------------------------------------------------------------------------#

#----------------------------------------------------------------------------#
# Busqueda de información
    
for x in range(0,len(rutas)):
    fila=1
    if opcion=='3' or opcion=='2':
        lista_archivos=os.listdir(rutas[x])
        
    # Crear archivo de salida
    Resultados=Workbook()
    Hoja=Resultados.active
    Hoja.title='Indices'
    Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))  
            
    # Buqueda en los archivos
    for recorrido in range(len(lista_archivos)):
        
        numero_atomos=0
        
        entrada_1=False 
        entrada_2=False
        
        # Busca archivos .log
        busqueda=re.compile(r'.*_N\.log')
        objeto=busqueda.search(lista_archivos[recorrido])
        if str(type(objeto)) != '<class \'NoneType\'>':
            entrada_1=True
        
        # Determina si terminaron adecuadamente y si existen
        if entrada_1==True:
            
            archivo1=objeto.group()
            archivo2=objeto.group().split('.')[0]+'-1'+'.log'
            archivo3=objeto.group().split('.')[0]+'+1'+'.log'
            triada_archivos=[archivo2,archivo1,archivo3]
         
            for k in range(0,3):             
                if os.path.exists(os.path.join(rutas[x],triada_archivos[k])):
                    for i in range(0,20):
                       archivo=open(os.path.join(rutas[x],triada_archivos[k]))
                       busqueda_normal=re.compile(r'Normal termination')
                       resultado=busqueda_normal.search(archivo.readlines()[-i])
                       if str(type(resultado)) != '<class \'NoneType\'>':
                           entrada_2=True                     
                           archivo.close()
                           break
                       archivo.close()
                       if entrada_2==False and i==19: 
                          archivos_malos.append(triada_archivos[k])

        # Busqueda en el archivo
        if entrada_2==True:
            
            Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
            Hoja=Resultados.active
            a0=Hoja.cell(row=fila,column=1,value=archivo1[:-6]+'.log')
            fila+=1
            Hoja.cell(row=fila,column=1,value='Atom number')
            Hoja.cell(row=fila,column=2,value='Symbol')
            Hoja.cell(row=fila,column=3,value='N-1')
            Hoja.cell(row=fila,column=4,value='N')
            Hoja.cell(row=fila,column=5,value='N+1')
            Hoja.cell(row=fila,column=6,value='CDD')
            Hoja.cell(row=fila,column=8,value='Atom number')
            Hoja.cell(row=fila,column=9,value='Symbol')
            Hoja.cell(row=fila,column=10,value='CDD')
            
            Hoja.cell(row=fila,column=12,value='Atom number')
            Hoja.cell(row=fila,column=13,value='Symbol')
            Hoja.cell(row=fila,column=14,value='N-1')
            Hoja.cell(row=fila,column=15,value='N')
            Hoja.cell(row=fila,column=16,value='N+1')
            Hoja.cell(row=fila,column=17,value='CDD')
            Hoja.cell(row=fila,column=19,value='Atom number')
            Hoja.cell(row=fila,column=20,value='Symbol')
            Hoja.cell(row=fila,column=21,value='CDD')
            
            Hoja.cell(row=fila,column=23,value='Fukui E+')
            Hoja.cell(row=fila,column=24,value='Fukui Nu-')
            Hoja.cell(row=fila,column=25,value='Fukui Radical')
            Hoja.cell(row=fila,column=27,value='Atom number')
            Hoja.cell(row=fila,column=28,value='Symbol')
            Hoja.cell(row=fila,column=29,value='Fukui E+')
            Hoja.cell(row=fila,column=31,value='Atom number')
            Hoja.cell(row=fila,column=32,value='Symbol')
            Hoja.cell(row=fila,column=33,value='Fukui Nu-')
            Hoja.cell(row=fila,column=35,value='Atom number')
            Hoja.cell(row=fila,column=36,value='Symbol')
            Hoja.cell(row=fila,column=37,value='Fukui Radical')            
            fila+=1
            Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
        
            # Numero de átomos
            archivo=open(os.path.join(rutas[x],archivo1))
            for k in range(0,len(archivo.readlines())):
                atomos_completos=False
                archivo.close()
                archivo=open(os.path.join(rutas[x],archivo1))
                busqueda_numero_atomos=re.compile(r'NAtoms=')
                if busqueda_numero_atomos.search(archivo.readlines()[k]) != None:
                    archivo.close()
                    archivo=open(os.path.join(rutas[x],objeto.group()))
                    numero_atomos=archivo.readlines()[k].split()
                    numero_atomos=int(numero_atomos[1])
                    archivo.close()
                    break
            
            # Añadir numeros a xlxs
            fila_n=fila
            for n in range(1,numero_atomos+1):
                Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                Hoja=Resultados.active
                Hoja.cell(row=fila_n,column=1,value=n)
                fila_n+=1
                Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                
            # Busqueda simbolos atomicos 
            simbolos = []
            archivo=open(os.path.join(rutas[x],archivo1))
            for k in range(0,len(archivo.readlines())):
                archivo=open(os.path.join(rutas[x],archivo1))
                busqueda_atomos=re.compile(r'^ Charge')
                if busqueda_atomos.search(archivo.readlines()[k]) != None:
                    archivo.close()
                    for z in range(1,numero_atomos+1):
                        archivo=open(os.path.join(rutas[x],archivo1))
                        atomo=archivo.readlines()[k+z].split()
                        simbolos.append(atomo[0])
                        archivo.close()
                    break
            
            # Añadir simbolos a archivos
            fila_n=fila
            for n in range(1,numero_atomos+1):
                Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                Hoja=Resultados.active
                Hoja.cell(row=fila_n,column=2,value=simbolos[n-1])
                fila_n+=1
                Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
            
            # Busqueda cargas de Hirshfeld
            fila_n=fila
            for h in range(0,3):
                archivo=open(os.path.join(rutas[x],triada_archivos[h]))
                for k in range(0,len(archivo.readlines())):
                    archivo=open(os.path.join(rutas[x],triada_archivos[h]))
                    busqueda=re.compile(r'^ Hirshfeld charges, spin densities')
                    if busqueda.search(archivo.readlines()[-k]) != None:
                        archivo.close()
                        for z in range(1,numero_atomos+1):
                            archivo=open(os.path.join(rutas[x],triada_archivos[h]))
                            carga=archivo.readlines()[(len(archivo.readlines())-k)+1+z].split()[7]
                            archivo.close()
                            Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                            Hoja=Resultados.active
                            Hoja.cell(row=fila_n,column=(h+3),value=float(carga))
                            Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                            fila_n+=1
                        break
                fila_n=fila
                
            # Carga de Hirshfeld con hidrógenos
            numero_atomos_pesados=0
            for i in range(0,len(simbolos)):
                if simbolos[i] != 'H':
                    numero_atomos_pesados= numero_atomos_pesados+1
            fila_n=fila
            for h in range(0,3):
                archivo=open(os.path.join(rutas[x],triada_archivos[h]))
                for k in range(0,len(archivo.readlines())):
                    archivo=open(os.path.join(rutas[x],triada_archivos[h]))
                    busqueda=re.compile(r'^ Hirshfeld charges (and spin densities )?with hydrogens summed into heavy atoms')
                    if busqueda.search(archivo.readlines()[-k]) != None:
                        archivo.close()
                        for z in range(1,numero_atomos_pesados+1):     
                            archivo=open(os.path.join(rutas[x],triada_archivos[h]))
                            if h != 1: 
                                 carga=archivo.readlines()[(len(archivo.readlines())-k)+1+z].split()[4]
                            else:
                                 carga=archivo.readlines()[(len(archivo.readlines())-k)+1+z].split()[3]      
                            archivo.close()
                            if h==0:     
                                archivo=open(os.path.join(rutas[x],triada_archivos[h]))
                                numero_atomo_pesado=archivo.readlines()[(len(archivo.readlines())-k)+1+z].split()[0]
                                archivo.close()
                                archivo=open(os.path.join(rutas[x],triada_archivos[h]))
                                simbolo_atomo_pesado=archivo.readlines()[(len(archivo.readlines())-k)+1+z].split()[1]
                                archivo.close()
                                Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                                Hoja=Resultados.active
                                Hoja.cell(row=fila_n,column=(h+12),value=int(numero_atomo_pesado))   
                                Hoja.cell(row=fila_n,column=(h+13),value=str(simbolo_atomo_pesado))
                                Resultados.save(os.path.join(rutas[x],'reactivity.xlsx')) 
                            Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                            Hoja=Resultados.active
                            Hoja.cell(row=fila_n,column=(h+14),value=float(carga))
                            Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                            fila_n+=1
                        break
                fila_n=fila
            
            # Cálculo CDD sin H
            for c in range(0,numero_atomos):
                Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                Hoja=Resultados.active
                a=Hoja.cell(row=fila_n,column=3).value
                b=Hoja.cell(row=fila_n,column=4).value
                c=Hoja.cell(row=fila_n,column=5).value
                Hoja.cell(row=fila_n,column=6,value=(round(2*b-c-a,5))) # Cálculo CDD sin H
                Hoja.cell(row=fila_n,column=23,value=(round(b-c,5))) # Cálculo Fukui E
                Hoja.cell(row=fila_n,column=24,value=(round(a-b,5))) # Cálculo Fukui Nu
                Hoja.cell(row=fila_n,column=25,value=(round((a-c)/2,5))) # Cálculo Fukui R
                fila_n+=1
                Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
              
             # Cálculo CDD con H
            fila_n=fila
            for c in range(0,numero_atomos_pesados):
                Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                Hoja=Resultados.active
                d=Hoja.cell(row=fila_n,column=14).value
                e=Hoja.cell(row=fila_n,column=15).value
                f=Hoja.cell(row=fila_n,column=16).value
                Hoja.cell(row=fila_n,column=17,value=(round(2*e-f-d,5))) # Cálculo CDD con H
                fila_n+=1
                Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))    
                
            # Ordenar datos CDD con H
            fila_n=fila
            orden=[]
            organizar=[]
            for o in range(0,numero_atomos):
                orden.append(Hoja.cell(row=fila_n,column=6).value)
                organizar.append(o)
                fila_n+=1
                
            orden=sorted(orden)

            fila_n=fila
            for o in range(0,numero_atomos):
                for k in organizar:
                    Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                    Hoja=Resultados.active
                    if orden[o]==Hoja.cell(row=fila_n+k,column=6).value:
                        Hoja.cell(row=fila_n+o,column=8,value=(Hoja.cell(row=fila_n+k,column=1).value))
                        Hoja.cell(row=fila_n+o,column=9,value=(Hoja.cell(row=fila_n+k,column=2).value))
                        Hoja.cell(row=fila_n+o,column=10,value=orden[o])
                        Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                        organizar.remove(k)
                        break
            
            # Ordenar datos Fukui+
            fila_n=fila
            orden=[]
            organizar=[]
            for o in range(0,numero_atomos):
                orden.append(Hoja.cell(row=fila_n,column=23).value)
                organizar.append(o)
                fila_n+=1
                
            orden=sorted(orden)

            fila_n=fila
            for o in range(0,numero_atomos):
                for k in organizar:
                    Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                    Hoja=Resultados.active
                    if orden[o]==Hoja.cell(row=fila_n+k,column=23).value:
                        Hoja.cell(row=fila_n+o,column=27,value=(Hoja.cell(row=fila_n+k,column=1).value))
                        Hoja.cell(row=fila_n+o,column=28,value=(Hoja.cell(row=fila_n+k,column=2).value))
                        Hoja.cell(row=fila_n+o,column=29,value=orden[o])
                        Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                        organizar.remove(k)
                        break
            
            # Ordenar datos fukui-
            fila_n=fila
            orden=[]
            organizar=[]
            for o in range(0,numero_atomos):
                orden.append(Hoja.cell(row=fila_n,column=24).value)
                organizar.append(o)
                fila_n+=1
                
            orden=sorted(orden)

            fila_n=fila
            for o in range(0,numero_atomos):
                for k in organizar:
                    Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                    Hoja=Resultados.active
                    if orden[o]==Hoja.cell(row=fila_n+k,column=24).value:
                        Hoja.cell(row=fila_n+o,column=31,value=(Hoja.cell(row=fila_n+k,column=1).value))
                        Hoja.cell(row=fila_n+o,column=32,value=(Hoja.cell(row=fila_n+k,column=2).value))
                        Hoja.cell(row=fila_n+o,column=33,value=orden[o])
                        Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                        organizar.remove(k)
                        break

            # Ordenar datos fukui radical
            fila_n=fila
            orden=[]
            organizar=[]
            for o in range(0,numero_atomos):
                orden.append(Hoja.cell(row=fila_n,column=25).value)
                organizar.append(o)
                fila_n+=1
                
            orden=sorted(orden)

            fila_n=fila
            for o in range(0,numero_atomos):
                for k in organizar:
                    Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                    Hoja=Resultados.active
                    if orden[o]==Hoja.cell(row=fila_n+k,column=25).value:
                        Hoja.cell(row=fila_n+o,column=35,value=(Hoja.cell(row=fila_n+k,column=1).value))
                        Hoja.cell(row=fila_n+o,column=36,value=(Hoja.cell(row=fila_n+k,column=2).value))
                        Hoja.cell(row=fila_n+o,column=37,value=orden[o])
                        Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                        organizar.remove(k)
                        break
            
             # Ordenar datos CDD con H
            fila_n=fila
            orden=[]
            organizar=[]
            for o in range(0,numero_atomos_pesados):
                orden.append(Hoja.cell(row=fila_n,column=17).value)
                organizar.append(o)
                fila_n+=1
                
            orden=sorted(orden)

            fila_n=fila
            for o in range(0,numero_atomos_pesados):
                for k in organizar:
                    Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                    Hoja=Resultados.active
                    if orden[o]==Hoja.cell(row=fila_n+k,column=17).value:
                        Hoja.cell(row=fila_n+o,column=19,value=(Hoja.cell(row=fila_n+k,column=12).value))
                        Hoja.cell(row=fila_n+o,column=20,value=(Hoja.cell(row=fila_n+k,column=13).value))
                        Hoja.cell(row=fila_n+o,column=21,value=orden[o])
                        Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                        organizar.remove(k)
                        break
        
            # Indice de Anderson
            # # Numeracion fila Nu
            for t in range(0,orden_reactividad):
                Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                Hoja=Resultados.active
                Hoja.cell(row=fila+t*(26)+numero_atomos+1,column=1,value='\u0394N/k Nu- [{0}]'.format(t+1))
                Hoja.cell(row=fila+t*(26)+numero_atomos+1,column=14,value='\u0394N/k Nu- [{0}]'.format(t+1))
                Resultados.save(os.path.join(rutas[x],'reactivity.xlsx')) 
                for a in range(0,11):
                  Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                  Hoja=Resultados.active
                  Hoja.cell(row=fila+t*(26)+numero_atomos+2+a,column=1,value=-(1-(a/10))) 
                  Hoja.cell(row=fila+t*(26)+numero_atomos+2+a,column=14,value=-(1-(a/10)))
                  Resultados.save(os.path.join(rutas[x],'reactivity.xlsx')) 
                  
                # # Numeración Columna
                for a in range(0,11):
                    Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                    Hoja=Resultados.active
                    Hoja.cell(row=fila+t*(26)+numero_atomos+1,column=2+a,value=(1-(a/5))) 
                    Hoja.cell(row=fila+t*(26)+numero_atomos+1,column=15+a,value=(1-(a/5)))
                    Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                
                # # Cálculo indice
                i,j = 0,0
                for k in numpy.arange(1,-1.2,-0.2):
                    for N in numpy.arange(-1,0.1,0.1):
                        nucleofilos=[]
                        for atomo in range(0,numero_atomos):
                             nu=(k+1+N*(k-1))*Hoja.cell(row=fila_n+atomo,column=4).value-N*(k-1)*Hoja.cell(row=fila_n+atomo,column=3).value
                             nucleofilos.append(nu)
                        nucleofilos_ordenados=sorted(nucleofilos)
                        atomo_mas_nucleofilico=nucleofilos.index(nucleofilos_ordenados[t])
                        Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                        Hoja=Resultados.active
                        Hoja.cell(row=fila+t*(26)+numero_atomos+j+2,column=2+i,value=nucleofilos_ordenados[t])
                        Hoja.cell(row=fila+t*(26)+numero_atomos+j+2,column=15+i,value=str(atomo_mas_nucleofilico+1)+' '+str(Hoja.cell(row=fila+atomo_mas_nucleofilico,column=2).value))
                        Resultados.save(os.path.join(rutas[x],'reactivity.xlsx')) 
                        j+=1
                    j=0
                    i+=1
    
                # # Numeracion fila E+
                Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                Hoja=Resultados.active
                Hoja.cell(row=fila+numero_atomos+t*(26)+14,column=1,value='\u0394N/k E+ [{0}]'.format(t+1))
                Hoja.cell(row=fila+numero_atomos+t*(26)+14,column=14,value='\u0394N/k E+ [{0}]'.format(t+1))
                Resultados.save(os.path.join(rutas[x],'reactivity.xlsx')) 
                for a in range(0,11):
                  Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                  Hoja=Resultados.active
                  Hoja.cell(row=fila+numero_atomos+t*(26)+15+a,column=1,value=(1-(a/10))) 
                  Hoja.cell(row=fila+numero_atomos+t*(26)+15+a,column=14,value=(1-(a/10)))
                  Resultados.save(os.path.join(rutas[x],'reactivity.xlsx')) 
                  
                # # Numeración Columna
                for a in range(0,11):
                    Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                    Hoja=Resultados.active
                    Hoja.cell(row=fila+numero_atomos+t*(26)+14,column=2+a,value=(1-(a/5)))
                    Hoja.cell(row=fila+numero_atomos+t*(26)+14,column=15+a,value=(1-(a/5))) 
                    Resultados.save(os.path.join(rutas[x],'reactivity.xlsx'))
                
                # # Cálculo indice
                i,j = 0,0
                for k in numpy.arange(1,-1.2,-0.2):
                    for N in numpy.arange(1,-0.1,-0.1):
                        electrofilos=[]
                        for atomo in range(0,numero_atomos):
                             e=(N*(k-1)-k-1)*Hoja.cell(row=fila_n+atomo,column=4).value-N*(k-1)*Hoja.cell(row=fila_n+atomo,column=5).value
                             electrofilos.append(e)
                        electrofilos_ordenados=sorted(electrofilos)
                        atomo_mas_electrofilico=electrofilos.index(electrofilos_ordenados[t])
                        Resultados = load_workbook(os.path.join(rutas[x],'reactivity.xlsx'))
                        Hoja=Resultados.active
                        Hoja.cell(row=fila+numero_atomos+t*(26)+14+j+1,column=2+i,value=electrofilos_ordenados[t])
                        Hoja.cell(row=fila+numero_atomos+t*(26)+14+j+1,column=15+i,value=str(atomo_mas_electrofilico+1)+' '+str(Hoja.cell(row=fila+atomo_mas_electrofilico,column=2).value))
                        Resultados.save(os.path.join(rutas[x],'reactivity.xlsx')) 
                        j+=1
                    j=0
                    i+=1
                
            #Archivos completados    
            print('Completado: ' + archivo1[:-6]+'.log')
            fila+=numero_atomos+2*(orden_reactividad)*13+1
                           
# Mostrar archivos que no terminarón correctamente           
if len(archivos_malos) != 0:
    if os.path.exists(os.path.join(rutas[x],'error.txt')):
            os.remove(os.path.join(rutas[x],'error.txt'))
    print('\nAbnormal termination: ',archivos_malos)
    archivo_salida=open(os.path.join(rutas[x],'error.txt'),'w')
    archivo_salida.write('Abnormal termination: \n')
    archivo_salida.close()
    for z in range(0,len(archivos_malos)):
        archivo_salida=open(os.path.join(rutas[x],'error.txt'),'a')
        archivo_salida.write('\n'+archivos_malos[z])
        archivo_salida.close()
        
#----------------------------------------------------------------------------#

